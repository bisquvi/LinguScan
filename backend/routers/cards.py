from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from database import get_db
import models
import schemas
from routers.auth import get_current_user
from services.srs_service import process_review

from typing import List
import datetime
import io
import random

router = APIRouter()


def get_deck_or_404(db: Session, deck_id: int, owner_id: int):
    deck = (
        db.query(models.Deck)
        .filter(models.Deck.id == deck_id, models.Deck.owner_id == owner_id)
        .first()
    )
    if not deck:
        raise HTTPException(status_code=404, detail="Deck not found")
    return deck


def get_card_or_404(db: Session, card_id: int, owner_id: int):
    card = (
        db.query(models.Card)
        .join(models.Deck, models.Card.deck_id == models.Deck.id)
        .filter(models.Card.id == card_id, models.Deck.owner_id == owner_id)
        .first()
    )
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    return card


def get_progress_or_404(db: Session, card_id: int, owner_id: int):
    progress = (
        db.query(models.Progress)
        .join(models.Card, models.Progress.card_id == models.Card.id)
        .join(models.Deck, models.Card.deck_id == models.Deck.id)
        .filter(models.Progress.card_id == card_id, models.Deck.owner_id == owner_id)
        .first()
    )
    if not progress:
        raise HTTPException(status_code=404, detail="Card progress not found")
    return progress


def get_quiz_session_or_404(db: Session, session_id: int, user_id: int):
    session = (
        db.query(models.QuizSession)
        .filter(models.QuizSession.id == session_id, models.QuizSession.user_id == user_id)
        .first()
    )
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return session


@router.post("/decks/", response_model=schemas.Deck)
def create_deck(
    deck: schemas.DeckCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    db_deck = models.Deck(name=deck.name, owner_id=current_user.id)
    db.add(db_deck)
    db.commit()
    db.refresh(db_deck)
    return db_deck


@router.get("/decks/", response_model=List[schemas.Deck])
def read_decks(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    return db.query(models.Deck).filter(models.Deck.owner_id == current_user.id).all()


@router.get("/decks/{deck_id}", response_model=schemas.Deck)
def read_deck(
    deck_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    return get_deck_or_404(db, deck_id, current_user.id)


@router.delete("/decks/{deck_id}", status_code=204)
def delete_deck(
    deck_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    deck = get_deck_or_404(db, deck_id, current_user.id)

    sessions_query = db.query(models.QuizSession.id).filter(models.QuizSession.deck_id == deck_id)
    db.query(models.QuizResult).filter(models.QuizResult.session_id.in_(sessions_query)).delete(synchronize_session=False)
    db.query(models.QuizSession).filter(models.QuizSession.deck_id == deck_id).delete(synchronize_session=False)

    db.delete(deck)
    db.commit()


@router.post("/decks/{deck_id}/cards", response_model=schemas.Card)
def create_card(
    deck_id: int,
    card: schemas.CardCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    get_deck_or_404(db, deck_id, current_user.id)

    db_card = models.Card(**card.model_dump(), deck_id=deck_id)
    db.add(db_card)
    db.commit()
    db.refresh(db_card)

    db_progress = models.Progress(card_id=db_card.id)
    db.add(db_progress)
    db.commit()
    return db_card


@router.post("/decks/{deck_id}/import-excel")
def import_excel(
    deck_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    get_deck_or_404(db, deck_id, current_user.id)

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    extension = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if extension not in ("xlsx", "xls"):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")

    try:
        from openpyxl import load_workbook

        contents = file.file.read()
        workbook = load_workbook(filename=io.BytesIO(contents), read_only=True)
        worksheet = workbook.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {exc}") from exc

    imported = 0
    skipped_header = False

    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if len(row) < 2:
            continue

        front = str(row[0]).strip() if row[0] else ""
        back = str(row[1]).strip() if row[1] else ""
        context = str(row[2]).strip() if len(row) > 2 and row[2] else None

        if not front or not back:
            continue

        if not skipped_header:
            skipped_header = True
            lower_front = front.lower()
            lower_back = back.lower()
            header_keywords = ["kelime", "word", "sözcük", "front", "ingilizce", "english"]
            back_keywords = ["anlam", "çeviri", "translation", "meaning", "back", "türkçe", "turkish"]
            if any(keyword in lower_front for keyword in header_keywords) or any(
                keyword in lower_back for keyword in back_keywords
            ):
                continue

        db_card = models.Card(front=front, back=back, context=context, deck_id=deck_id)
        db.add(db_card)
        db.flush()
        db.add(models.Progress(card_id=db_card.id))
        imported += 1

    db.commit()
    workbook.close()

    return {"imported": imported, "deck_id": deck_id}


@router.delete("/cards/{card_id}", status_code=204)
def delete_card(
    card_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    card = get_card_or_404(db, card_id, current_user.id)
    db.query(models.QuizResult).filter(models.QuizResult.card_id == card_id).delete(synchronize_session=False)
    db.delete(card)
    db.commit()


@router.post("/cards/{card_id}/review")
def review_card_legacy(
    card_id: int,
    quality: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    progress = get_progress_or_404(db, card_id, current_user.id)
    rating_map = {0: "again", 1: "again", 2: "mid", 3: "mid", 4: "good", 5: "easy"}
    rating = rating_map.get(quality, "mid")
    updates = process_review(progress, rating, is_correct=(quality >= 3))

    for key, value in updates.items():
        setattr(progress, key, value)

    db.commit()
    db.refresh(progress)
    return progress


@router.post("/quiz/session/start")
def start_quiz_session(
    payload: schemas.QuizSessionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    deck = get_deck_or_404(db, payload.deck_id, current_user.id)
    if not deck.cards:
        raise HTTPException(status_code=400, detail="No cards in this deck")

    session = models.QuizSession(
        deck_id=payload.deck_id,
        user_id=current_user.id,
        quiz_type=payload.quiz_type,
        total_cards=len(deck.cards),
    )
    db.add(session)
    db.commit()
    db.refresh(session)

    cards = deck.cards

    if payload.quiz_type == "multiple_choice":
        all_cards = (
            db.query(models.Card)
            .join(models.Deck, models.Card.deck_id == models.Deck.id)
            .filter(models.Deck.owner_id == current_user.id)
            .all()
        )
        all_backs = list({card.back for card in all_cards})

        quiz_cards = []
        for card in cards:
            distractors = [back for back in all_backs if back != card.back]
            chosen = random.sample(distractors, 3) if len(distractors) >= 3 else distractors + ["—"] * (3 - len(distractors))

            options = chosen + [card.back]
            random.shuffle(options)

            quiz_cards.append(
                {
                    "card_id": card.id,
                    "question": card.front,
                    "options": options,
                    "correct_index": options.index(card.back),
                }
            )

        return {"session_id": session.id, "quiz_type": "multiple_choice", "cards": quiz_cards}

    quiz_cards = [{"card_id": card.id, "prompt": card.front, "answer": card.back} for card in cards]
    return {"session_id": session.id, "quiz_type": "recall", "cards": quiz_cards}


@router.post("/quiz/session/{session_id}/review/{card_id}")
def review_quiz_card(
    session_id: int,
    card_id: int,
    payload: schemas.QuizReviewRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    session = get_quiz_session_or_404(db, session_id, current_user.id)

    card = db.query(models.Card).filter(models.Card.id == card_id, models.Card.deck_id == session.deck_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Card not found in this quiz session")

    progress = db.query(models.Progress).filter(models.Progress.card_id == card_id).first()
    if not progress:
        raise HTTPException(status_code=404, detail="Card progress not found")

    updates = process_review(progress, payload.rating, payload.is_correct)
    for key, value in updates.items():
        setattr(progress, key, value)

    db.add(
        models.QuizResult(
            session_id=session_id,
            card_id=card_id,
            is_correct=1 if payload.is_correct else 0,
            rating=payload.rating,
        )
    )
    db.commit()

    return {"status": "ok", "level": updates["level"], "next_review": updates["next_review"]}


@router.post("/quiz/session/{session_id}/finish")
def finish_quiz_session(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    session = get_quiz_session_or_404(db, session_id, current_user.id)
    results = db.query(models.QuizResult).filter(models.QuizResult.session_id == session_id).all()

    correct = sum(result.is_correct for result in results)
    wrong = len(results) - correct
    accuracy = round(correct / len(results) * 100, 1) if results else 0.0

    session.finished_at = datetime.datetime.utcnow()
    session.correct_answers = correct
    session.wrong_answers = wrong
    session.accuracy = accuracy
    db.commit()

    reviewed_card_ids = [result.card_id for result in results]
    hardest_progresses = (
        db.query(models.Progress)
        .filter(models.Progress.card_id.in_(reviewed_card_ids))
        .order_by(models.Progress.difficulty_score.desc())
        .limit(3)
        .all()
    )
    easiest_progresses = (
        db.query(models.Progress)
        .filter(models.Progress.card_id.in_(reviewed_card_ids))
        .order_by(models.Progress.difficulty_score.asc())
        .limit(3)
        .all()
    )

    hardest_cards = [db.get(models.Card, progress.card_id) for progress in hardest_progresses]
    easiest_cards = [db.get(models.Card, progress.card_id) for progress in easiest_progresses]

    return {
        "session_id": session_id,
        "deck_id": session.deck_id,
        "quiz_type": session.quiz_type,
        "total_cards": session.total_cards,
        "correct_answers": correct,
        "wrong_answers": wrong,
        "accuracy": accuracy,
        "hardest_cards": [{"id": card.id, "front": card.front, "back": card.back} for card in hardest_cards if card],
        "easiest_cards": [{"id": card.id, "front": card.front, "back": card.back} for card in easiest_cards if card],
    }


@router.get("/dashboard")
def get_dashboard(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    all_progresses = (
        db.query(models.Progress)
        .join(models.Card, models.Progress.card_id == models.Card.id)
        .join(models.Deck, models.Card.deck_id == models.Deck.id)
        .filter(models.Deck.owner_id == current_user.id)
        .all()
    )

    total_cards = len(all_progresses)
    new_cards = sum(1 for progress in all_progresses if progress.level == "NEW")
    learning_cards = sum(1 for progress in all_progresses if progress.level == "LEARNING")
    review_cards = sum(1 for progress in all_progresses if progress.level == "REVIEW")
    mastered_cards = sum(1 for progress in all_progresses if progress.level == "MASTERED")

    total_quizzes = db.query(models.QuizSession).filter(models.QuizSession.user_id == current_user.id).count()
    rates = [progress.success_rate for progress in all_progresses if progress.total_reviews > 0]
    overall_success_rate = round(sum(rates) / len(rates) * 100, 1) if rates else 0.0

    difficult = (
        db.query(models.Progress)
        .join(models.Card, models.Progress.card_id == models.Card.id)
        .join(models.Deck, models.Card.deck_id == models.Deck.id)
        .filter(models.Progress.total_reviews > 0, models.Deck.owner_id == current_user.id)
        .order_by(models.Progress.difficulty_score.desc())
        .limit(5)
        .all()
    )
    easiest = (
        db.query(models.Progress)
        .join(models.Card, models.Progress.card_id == models.Card.id)
        .join(models.Deck, models.Card.deck_id == models.Deck.id)
        .filter(models.Progress.total_reviews > 0, models.Deck.owner_id == current_user.id)
        .order_by(models.Progress.difficulty_score.asc())
        .limit(5)
        .all()
    )

    def card_summary(progress: models.Progress):
        card = db.get(models.Card, progress.card_id)
        return {"id": card.id, "front": card.front, "back": card.back} if card else None

    difficult_cards = [summary for summary in (card_summary(progress) for progress in difficult) if summary]
    easiest_cards = [summary for summary in (card_summary(progress) for progress in easiest) if summary]

    return {
        "total_cards": total_cards,
        "new_cards": new_cards,
        "learning_cards": learning_cards,
        "review_cards": review_cards,
        "mastered_cards": mastered_cards,
        "total_quizzes": total_quizzes,
        "overall_success_rate": overall_success_rate,
        "most_difficult_cards": difficult_cards,
        "best_known_cards": easiest_cards,
    }
